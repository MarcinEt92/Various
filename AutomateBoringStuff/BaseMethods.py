import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import pywhatkit


class BaseMethods:
    APPLICATION_PATH = os.path.dirname(sys.executable)
    CHROME_DRIVER_PATH = "data/chrome_driver"
    WEB_ADDRESS = "https://www.thesun.co.uk/sport/football/"

    datetime_now = datetime.now()
    datetime_now_str = datetime_now.strftime("%m%d%Y")

    @classmethod
    def get_simpson_html_data_with_pandas(cls):
        web_address = "https://en.wikipedia.org/wiki/List_of_The_Simpsons_episodes_(seasons_1%E2%80%9320)"
        sim_list = pd.read_html(web_address)  # read_html returns a list
        season_1_index = 1
        first_season = sim_list[season_1_index]

        print(f"There are {len(sim_list)} tables on web site")
        print(f"\nSimpsons 1st season: \n{first_season}")

    @classmethod
    def get_csv_data_from_website(cls):
        web_address = "https://datahub.io/sports-data/english-premier-league/r/season-1819.csv"
        csv_data = pd.read_csv(web_address)

        # renaming column header:
        csv_data.rename(columns={"Div": "Division"}, inplace=True)

        # print data:
        print(f"Csv data: \n{csv_data}")

    @classmethod
    def extract_table_from_pdf(cls):
        # camelot-py lib required
        # web_address = "https://www.w3.org/WAI/WCAG21/working-examples/pdf-table/table.pdf"
        # pdf_path = "data/tab.pdf"

        # commented because they do not work properly
        # tables = camelot.read_pdf(pdf_path, pages=1)
        # tables.export("tab.csv", f="csv", compress=True)
        # tables[0].to_csv("tab0.csv")
        pass

    @classmethod
    def web_automation_scraping_html_tags(cls):
        # XPath syntax:
        # //tagName
        # //tagName[1]
        # //tagName[@AttributeName="Value"]
        # //tagName[contains(@AttributeName, "Value")]
        # and or operators are being used between expressions
        # https://scrapinghub.github.io/xpath-playground/
        # //h2/a or //h2/text()
        # /: Select the children from the node set
        # //: matching node set should be located at any level within document
        # //tagName/* - all elements inside a tag
        pass

    @classmethod
    def web_automation(cls):
        service = Service(executable_path=cls.CHROME_DRIVER_PATH)
        driver = webdriver.Chrome(service=service)
        cls.export_data_to_csv(driver)
        driver.quit()

    @classmethod
    def web_automation_headless_mode(cls):
        options = Options()
        options.headless = True
        service = Service(executable_path=cls.CHROME_DRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        cls.export_data_to_csv(driver)
        driver.quit()

    @classmethod
    def export_data_to_csv(cls, driver):
        driver.get(cls.WEB_ADDRESS)

        # find elements in the news:
        containers = driver.find_elements(by='xpath', value='//div[@class="teaser__copy-container"]')

        titles = []
        subtitles = []
        links = []

        for container in containers:
            # now we have to provide reference value only
            title = container.find_element(by='xpath', value='./a/h3').text
            subtitle = container.find_element(by='xpath', value='./a/p').text
            link = container.find_element(by='xpath', value='./a').get_attribute('href')

            # if len(title) > 0:
            titles.append(title)
            subtitles.append(subtitle)
            links.append(link)

        # export data to csv, create dataframe using pandas
        # football_data_csv_path = f"data/football_news/headlines_{cls.datetime_now_str}.csv"
        file_name = f"headlines_{cls.datetime_now_str}.csv"
        football_data_csv_path = os.path.join(os.getcwd(), "data", file_name)
        data_dict = {'titles': titles, 'subtitles': subtitles, 'links': links}
        df_headlines = pd.DataFrame(data_dict)
        df_headlines.to_csv(football_data_csv_path)

    @classmethod
    def pivot_tables_excel(cls):
        excel_path = "data/sales.xlsx"
        df = pd.read_excel(excel_path)

        # select multiple columns from excel file
        df = df[['Gender', 'Product line', 'Total']]

        # index, columns, values, operation to apply
        # we have now information how much each gender spent on various product lines
        pivot_table = df.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum')
        pivot_table.to_excel("data/sales_pivot.xlsx", "Report", startrow=4)
        print(pivot_table)

    @classmethod
    def excel_add_bar_chart(cls):
        work_book = load_workbook("data/sales_pivot.xlsx")
        sheet = work_book['Report']

        # setting range of our table
        min_column = work_book.active.min_column
        max_column = work_book.active.max_column
        min_row = work_book.active.min_row
        max_row = work_book.active.max_row

        bar_chart = BarChart()

        data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row+1, max_row=max_row)
        categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        sheet.add_chart(bar_chart, "B12")

        bar_chart.title = "sales by Product line"
        bar_chart.style = 5
        work_book.save("data/barchart.xlsx")

        print(min_row)

    @classmethod
    def automate_whatsapp(cls):
        pywhatkit.sendwhatmsg("+48515633245", "Test", 1, 5)
